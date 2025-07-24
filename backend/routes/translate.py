# Translate text using GPT
from flask import Blueprint, request, jsonify, send_file
from services.firestore_client import db
from services.openai_client import client as openai_client
from utils.char_counter import estimate_chars_in_file
from utils.excel_converter import convert_xls_to_xlsx
from openpyxl import load_workbook
from io import BytesIO
import os
import tempfile
import time

translate_bp = Blueprint("translate", __name__)


@translate_bp.route("/translate", methods=["POST"])
def translate_excel():
    uid = request.headers.get("X-User-Id")
    uploaded_file = request.files["file"]

    if not uid or not uploaded_file:
        return jsonify({"error": "Missing user ID or uploaded file"}), 400

    source_lang = request.form.get("sourceLang", "auto")
    target_lang = request.form.get("targetLang", "en")

    user_ref = db.collection("users").document(uid)
    user_doc = user_ref.get()

    if not user_doc.exists:
        return jsonify({"error": "User document not found."}), 404

    user_data = user_doc.to_dict()

    # Get user data
    char_used = user_data.get("charUsed", 0)
    char_limit = user_data.get("charLimit", 0)

    # Estimate new characters
    new_chars = estimate_chars_in_file(uploaded_file)
    if char_used + new_chars > char_limit:
        return jsonify({"error": "Character limit exceeded"}), 403

    filename = uploaded_file.filename.lower()
    start_time = time.time()
    translation_cache = {}

    try:
        if filename.endswith(".xls"):
            filepath = convert_xls_to_xlsx(uploaded_file)
        else:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_xlsx:
                filepath = temp_xlsx.name
                uploaded_file.save(filepath)

        wb = load_workbook(filepath)

        for ws in wb.worksheets:  # 🔄 Loop over all sheets
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        original = cell.value.strip()
                        if original in translation_cache:
                            translated = translation_cache[original]
                        else:
                            translated = translate_text(
                                original, source_lang, target_lang
                            )
                            translation_cache[original] = translated
                        cell.value = translated

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Update usage
        user_ref.update({"charUsed": char_used + new_chars})

        duration = time.time() - start_time  # ⏱ End timer
        print(f"[DEBUG] Translation completed in {duration:.2f} seconds")

        original_name = os.path.splitext(uploaded_file.filename)[0]
        translated_name = translate_text(original_name, source_lang, target_lang)
        final_filename = f"{original_name} ({translated_name}).xlsx"
        print("[INFO] final_filename:", final_filename)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=final_filename,
        )

    except RuntimeError as err:
        return jsonify({"error": str(err)}), 400


def translate_text(text, source_lang, target_lang):
    prompt = (
        f"You are a professional translator. "
        f"Translate the following text from {source_lang} to {target_lang} as accurately as possible, "
        f"preserving its contextual meaning. If the text is a number, symbol, or foreign-language string "
        f"that does not need translation, return it as is without modification. "
        f"Do not explain or add anything. Return only the translated result.\n\n"
        f"⚠️ Do not save, store, log, or use any part of this content for any reason. "
        f"All text is confidential and must not be retained after this operation.\n\n"
        f"Text:\n{text}"
    )
    try:
        response = openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"[ERROR] Failed to translate text:\n{e}")
        return text
