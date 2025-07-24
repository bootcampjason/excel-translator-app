from openpyxl import load_workbook
from io import BytesIO


def estimate_chars_in_file(file_storage):
    in_memory_file = BytesIO(file_storage.read())
    wb = load_workbook(in_memory_file, read_only=True)
    total_chars = 0

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    total_chars += len(cell)

    file_storage.seek(0)  # reset for re-use
    return total_chars
