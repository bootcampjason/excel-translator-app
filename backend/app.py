from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from io import BytesIO
from dotenv import load_dotenv
from openai import OpenAI
from firebase_admin import credentials, firestore, initialize_app
from utils.char_counter import estimate_chars_in_file
import os
import xlwings as xw
import tempfile
import time
from datetime import datetime, timezone, timedelta
import stripe
import json

# Load environment variables
load_dotenv()

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

app = Flask(__name__)
CORS(app, expose_headers=["Content-Disposition"])

# Firebase Admin Init
firebase_json = os.getenv("FIREBASE_CONFIG_JSON")
if not firebase_json:
    raise RuntimeError("FIREBASE_CONFIG_JSON environment variable is missing.")

firebase_creds = json.loads(firebase_json)
# cred = credentials.Certificate(os.getenv("GOOGLE_APPLICATION_CREDENTIALS"))
cred = credentials.Certificate(firebase_creds)
initialize_app(cred)
db = firestore.client()

PLANS = {"free": 10000, "starter": 50000, "pro": 200000}


@app.route("/")
def home():
    return "✅ Backend is running!"


def get_user_data(uid):
    if not uid:
        return None, jsonify({"error": "Missing user ID"}), 400

    user_ref = db.collection("users").document(uid)
    user_doc = user_ref.get()

    if not user_doc.exists:
        print(f"[DEBUG] User uid not found. Creating default user record.")
        user_data = {
            "plan": "free",
            "charUsedThisMonth": 0,
            "lastReset": firestore.SERVER_TIMESTAMP,
        }
        user_ref.set(user_data)
        return user_data

    user_data = user_doc.to_dict()
    last_reset = user_data.get("lastReset")

    if should_reset_monthly_usage(last_reset):
        print(f"[DEBUG] Resetting monthly usage for user")
        user_data["charUsedThisMonth"] = 0
        user_data["lastReset"] = firestore.SERVER_TIMESTAMP
        user_ref.update(
            {"charUsedThisMonth": 0, "lastReset": firestore.SERVER_TIMESTAMP}
        )

    return user_data


def should_reset_monthly_usage(last_reset):
    """Return True if last_reset was more than 30 days ago or missing."""
    if not last_reset:
        return True
    if isinstance(last_reset, datetime):
        last_reset = last_reset.replace(tzinfo=timezone.utc)
    return (datetime.now(timezone.utc) - last_reset) > timedelta(days=30)


@app.route("/usage", methods=["GET"])
def get_usage():
    uid = request.headers.get("X-User-Id")
    user_data, error_response, status = get_user_data(uid)
    if error_response:
        return error_response, status

    plan = user_data.get("plan", "free")
    char_used = user_data.get("charUsedThisMonth", 0)
    char_limit = PLANS.get(plan, 10000)

    return jsonify({"plan": plan, "charUsed": char_used, "charLimit": char_limit})


@app.route("/translate", methods=["POST"])
def translate_excel():
    uid = request.headers.get("X-User-Id")
    if not uid:
        return jsonify({"error": "Missing user ID"}), 400

    uploaded_file = request.files["file"]
    source_lang = request.form.get("sourceLang", "auto")
    target_lang = request.form.get("targetLang", "en")

    # Get user data (with reset logic)
    user_data = get_user_data(uid)
    plan = user_data.get("plan", "free")
    char_limit = PLANS.get(plan, 10000)
    char_used = user_data.get("charUsedThisMonth", 0)

    # Estimate new characters
    new_chars = estimate_chars_in_file(uploaded_file)
    if char_used + new_chars > char_limit:
        return jsonify({"error": "Character limit exceeded"}), 403

    filename = uploaded_file.filename.lower()
    start_time = time.time()
    translation_cache = {}

    try:
        if filename.endswith(".xls"):
            filepath = convert_xls_to_xlsx(uploaded_file)
        else:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_xlsx:
                filepath = temp_xlsx.name
                uploaded_file.save(filepath)

        wb = load_workbook(filepath)

        for ws in wb.worksheets:  # 🔄 Loop over all sheets
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip():
                        original = cell.value.strip()
                        if original in translation_cache:
                            translated = translation_cache[original]
                        else:
                            translated = translate_text(
                                original, source_lang, target_lang
                            )
                            translation_cache[original] = translated
                        cell.value = translated

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Update usage
        db.collection("users").document(uid).update(
            {"charUsedThisMonth": char_used + new_chars}
        )

        duration = time.time() - start_time  # ⏱ End timer
        print(f"[DEBUG] Translation completed in {duration:.2f} seconds")

        original_name = os.path.splitext(uploaded_file.filename)[0]
        translated_name = translate_text(original_name, source_lang, target_lang)
        final_filename = f"{original_name} ({translated_name}).xlsx"
        print("final_filename", final_filename)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=final_filename,
        )

    except RuntimeError as err:
        return jsonify({"error": str(err)}), 400


# Convert .xls to .xlsx using Excel via xlwings
def convert_xls_to_xlsx(xls_file):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xls") as temp_xls:
        xls_path = temp_xls.name
        xls_file.save(xls_path)

    xlsx_path = xls_path.replace(".xls", ".xlsx")

    try:
        app_excel = xw.App(visible=False)
        wb = app_excel.books.open(xls_path)
        wb.save(xlsx_path)
        wb.close()
        app_excel.quit()
        return xlsx_path
    except Exception as e:
        print(f"[ERROR] Excel not available: {e}")
        raise RuntimeError(
            "Microsoft Excel is not installed or accessible. Please upload .xlsx files only."
        )


# Translate text using GPT
def translate_text(text, source_lang, target_lang):
    prompt = (
        f"You are a professional translator. "
        f"Translate the following text from {source_lang} to {target_lang} as accurately as possible, "
        f"preserving its contextual meaning. If the text is a number, symbol, or foreign-language string "
        f"that does not need translation, return it as is without modification. "
        f"Do not explain or add anything. Return only the translated result.\n\n"
        f"⚠️ Do not save, store, log, or use any part of this content for any reason. "
        f"All text is confidential and must not be retained after this operation.\n\n"
        f"Text:\n{text}"
    )
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"[ERROR] Failed to translate text:\n{e}")
        return text


stripe.api_key = os.getenv("STRIPE_SECRET_KEY")

@app.route("/create-checkout-session", methods=["POST"])
def create_checkout_session():
    try:
        data = request.get_json()
        uid = data.get("uid")
        price_id = data.get("priceId")  # Retrieved from Stripe Dashboard
        if not uid or not price_id:
            return jsonify({"error": "Missing uid or priceId"}), 400

        session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            mode="subscription",
            line_items=[{"price": price_id, "quantity": 1}],
            success_url="http://localhost:3000/payment-success",
            cancel_url="http://localhost:3000/payment-cancel",
            metadata={"uid": uid},
        )

        return jsonify({"url": session.url})

    except stripe.error.StripeError as e:
        print(f"[STRIPE ERROR] {e}")
        return jsonify({"error": "A payment error occurred. Please try again."}), 500

    except Exception as e:
        print(f"[SERVER ERROR] Unexpected error during checkout session creation: {e}")
        return (
            jsonify({"error": "An unexpected error occurred. Please try again."}),
            500,
        )


@app.route("/ping", methods=["GET"])
def ping():
    return {"message": "Backend is alive!"}


if __name__ == "__main__":
    app.run(debug=True)
